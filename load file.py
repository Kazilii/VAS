from openpyxl import load_workbook

wb2 = load_workbook('sameple.xlsx')

print(wb2.sheetnames)

my_range = wb2.defined_names['Sheet']

dests = my_range.destinations

cells = []

for title, coord in dests:
    ws = wb2[title]
    cells.append

print(cells)